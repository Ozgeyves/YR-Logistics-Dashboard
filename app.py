import streamlit as st
import pandas as pd
import numpy as np
import re
import base64
import requests
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

st.set_page_config(page_title="YR Logistics Dashboard", layout="wide")

APP_DIR = Path(__file__).parent
HISTORY_DIR = APP_DIR / "gecmis_raporlar"
HISTORY_DIR.mkdir(exist_ok=True)


APP_NAME = "YR Logistics Dashboard"
APP_VERSION = "1.1.0"


def get_secret(name, default=""):
    try:
        return str(st.secrets.get(name, default))
    except Exception:
        return default


def show_app_header(subtitle):
    logo_path = APP_DIR / "logo.png"
    if logo_path.exists():
        logo_col, title_col = st.columns([1, 8])
        with logo_col:
            st.image(str(logo_path), width=90)
        with title_col:
            st.title(APP_NAME)
            st.caption(subtitle)
    else:
        st.title(APP_NAME)
        st.caption(subtitle)


def iso_week_label(value):
    """Tarihi ISO hafta formatında (örn. 2026-W29) döndürür."""
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return ""
    iso = ts.isocalendar()
    return f"{int(iso.year)}-W{int(iso.week):02d}"


def filter_report_names(report_names, query):
    query = str(query or "").strip().lower()
    if not query:
        return report_names
    return [name for name in report_names if query in name.lower()]

def _github_report_settings():
    """
    Streamlit Secrets varsa GitHub reports klasörünü kullanır.
    Yoksa local gecmis_raporlar klasörüne düşer.
    """
    try:
        token = str(st.secrets["GITHUB_TOKEN"]).strip()
        repo = str(st.secrets["GITHUB_REPO"]).strip()
        branch = str(st.secrets.get("GITHUB_BRANCH", "main")).strip()
        folder = str(st.secrets.get("GITHUB_REPORTS_FOLDER", "reports")).strip().strip("/")
        if token and repo:
            return token, repo, branch, folder
    except Exception:
        pass
    return None


def _github_headers(token):
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def list_saved_reports():
    settings = _github_report_settings()

    if settings is None:
        return sorted([p.name for p in HISTORY_DIR.glob("*.xlsx")], reverse=True)

    token, repo, branch, folder = settings
    url = f"https://api.github.com/repos/{repo}/contents/{folder}"

    response = requests.get(
        url,
        headers=_github_headers(token),
        params={"ref": branch},
        timeout=30,
    )

    if response.status_code == 404:
        return []

    if not response.ok:
        raise RuntimeError(f"GitHub rapor listesi okunamadı: {response.status_code} - {response.text[:200]}")

    reports = [
        item["name"]
        for item in response.json()
        if item.get("type") == "file"
        and str(item.get("name", "")).lower().endswith(".xlsx")
    ]

    return sorted(reports, reverse=True)


def load_saved_report(filename):
    settings = _github_report_settings()

    if settings is None:
        return (HISTORY_DIR / filename).read_bytes()

    token, repo, branch, folder = settings
    url = f"https://api.github.com/repos/{repo}/contents/{folder}/{filename}"

    response = requests.get(
        url,
        headers=_github_headers(token),
        params={"ref": branch},
        timeout=30,
    )
    if not response.ok:
        raise RuntimeError(f"GitHub raporu açılamadı: {response.status_code} - {response.text[:200]}")

    return base64.b64decode(response.json()["content"])


def save_report_permanently(filename, report_bytes):
    """
    GitHub ayarlıysa reports/ klasörüne kalıcı kaydeder.
    Aynı isim varsa otomatik Rev2, Rev3 şeklinde yeni isim oluşturur.
    """
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    settings = _github_report_settings()

    if settings is None:
        target = HISTORY_DIR / filename

        if target.exists():
            stem = target.stem
            suffix = target.suffix
            revision = 2
            while target.exists():
                target = HISTORY_DIR / f"{stem}_Rev{revision}{suffix}"
                revision += 1

        target.write_bytes(report_bytes)
        return target.name

    token, repo, branch, folder = settings

    original_stem = Path(filename).stem
    suffix = ".xlsx"
    candidate = filename
    revision = 2

    while True:
        path = f"{folder}/{candidate}"
        url = f"https://api.github.com/repos/{repo}/contents/{path}"

        check = requests.get(
            url,
            headers=_github_headers(token),
            params={"ref": branch},
            timeout=30,
        )

        if check.status_code == 404:
            break
        if check.status_code != 200:
            check.raise_for_status()

        candidate = f"{original_stem}_Rev{revision}{suffix}"
        revision += 1

    payload = {
        "message": f"Add YR Logistics Dashboard report: {candidate}",
        "content": base64.b64encode(report_bytes).decode("ascii"),
        "branch": branch,
    }

    response = requests.put(
        url,
        headers=_github_headers(token),
        json=payload,
        timeout=60,
    )
    if not response.ok:
        raise RuntimeError(f"GitHub rapor kaydı başarısız: {response.status_code} - {response.text[:300]}")

    return candidate

# ------------------------------------------------------------
# GİRİŞ / ROL SEÇİMİ
# ------------------------------------------------------------
if "role" not in st.session_state:
    st.session_state["role"] = None

if st.session_state["role"] is None:
    show_app_header("Giriş için şifrenizi yazın.")

    password = st.text_input("Şifre", type="password")

    if st.button("Giriş Yap", type="primary"):
        tedarik_password = get_secret("TEDARIK_PASSWORD", "tedarik")
        depo_password = get_secret("DEPO_PASSWORD", "depo")

        if password == tedarik_password:
            st.session_state["role"] = "tedarik"
            st.rerun()
        elif password == depo_password:
            st.session_state["role"] = "depo"
            st.rerun()
        else:
            st.error("Şifre hatalı.")

    st.stop()


# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def excel_serial_to_date(value):
    """Excel serial veya normal tarih değerini pandas datetime'a çevirir."""
    if pd.isna(value):
        return pd.NaT

    # Excel serial date
    if isinstance(value, (int, float, np.integer, np.floating)):
        return pd.to_datetime("1899-12-30") + pd.to_timedelta(float(value), unit="D")

    # String / datetime
    return pd.to_datetime(value, dayfirst=True, errors="coerce")


def normalize_code(value):
    """Ürün kodlarını 504.0 / '00504' / '4188 UN1266' gibi durumlara karşı normalize eder."""
    if pd.isna(value):
        return None

    text = str(value).strip()

    # 4188 UN1266 gibi hücrelerde ilk sayı ürün kodudur.
    match = re.search(r"\d+", text)
    if match:
        text = match.group(0)

    if text.endswith(".0"):
        text = text[:-2]

    return text.lstrip("0") or "0"


def clean_text(value):
    """Türkçe karakter / boşluk / büyük-küçük harf farklarını temizler."""
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()

    replacements = {
        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "ü": "u",
        "ş": "s",
        "ö": "o",
        "ç": "c",
        "â": "a",
        "î": "i",
        "û": "u",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ").replace("_", " ").replace("/", " ")
    text = " ".join(text.split())

    return text


def normalize_product_type(value):
    """
    Ürün tipi dosyasındaki farklı yazımları tek standarda indirir.
    Sadece 3 kategori kullanıyoruz:
    - Ana Ürün
    - Mini Sample
    - ADR, ADR sheetinden ayrıca yakalanır.
    """
    text = clean_text(value)

    if not text:
        return "Ana Ürün"

    # Önce Mini/Sample yakalanmalı. Çünkü bazı açıklamalarda ana kelimesi de geçebilir.
    mini_keywords = ["mini", "sample", "tester", "sachet", "numune", "deneme", "mini sample"]
    if any(k in text for k in mini_keywords):
        return "Mini Sample"

    ana_keywords = ["ana", "main", "regular", "standart", "standard"]
    if any(k in text for k in ana_keywords):
        return "Ana Ürün"

    # Tanınmayanları kaybetmemek için Ana Ürün kabul ediyoruz.
    return "Ana Ürün"


def get_week_start(date_series):
    """Her tarihi haftanın pazartesi gününe çeker."""
    date_series = pd.to_datetime(date_series, errors="coerce")
    return date_series - pd.to_timedelta(date_series.dt.weekday, unit="D")


def read_code_sheet(xls, possible_sheet_names):
    """
    Verilen olası sheet isimlerinden birini bulur.
    Sheet içindeki ilk dolu kolonu ürün kodu kabul eder.
    """
    existing_sheets = {clean_text(s): s for s in xls.sheet_names}

    selected_sheet = None
    for name in possible_sheet_names:
        key = clean_text(name)
        if key in existing_sheets:
            selected_sheet = existing_sheets[key]
            break

    if selected_sheet is None:
        return set(), None

    df = pd.read_excel(xls, sheet_name=selected_sheet)
    df = df.dropna(how="all")

    if df.empty:
        return set(), selected_sheet

    # İlk dolu kolonu kod kolonu kabul et
    first_valid_col = None
    for col in df.columns:
        if df[col].notna().sum() > 0:
            first_valid_col = col
            break

    if first_valid_col is None:
        return set(), selected_sheet

    codes = set(df[first_valid_col].apply(normalize_code).dropna().astype(str))
    return codes, selected_sheet



def read_campaign_sheet(xls):
    """
    Kampanya sheetini esnek okur.
    Desteklenen yapı örnekleri:
    - Sheet adı kampanya / Kampanya / campaign
    - Kolonlar: WINDOW-Campaign / Start / End
    - Header satırı dosyada 1., 2. veya 3. satırda olabilir.
    """
    campaign_sheet = None

    for sheet in xls.sheet_names:
        sheet_key = clean_text(sheet)
        if "kampanya" in sheet_key or "campaign" in sheet_key:
            campaign_sheet = sheet
            break

    if campaign_sheet is None:
        return pd.DataFrame(columns=["campaign", "start", "end"])

    raw = pd.read_excel(xls, sheet_name=campaign_sheet, header=None)
    raw = raw.dropna(how="all")

    if raw.empty:
        return pd.DataFrame(columns=["campaign", "start", "end"])

    header_row = None
    campaign_col = None
    start_col = None
    end_col = None

    # İlk 15 satırda header arıyoruz
    for idx in range(min(15, len(raw))):
        row = raw.iloc[idx]

        for col_idx, value in row.items():
            text = clean_text(value)

            if not text:
                continue

            if campaign_col is None and (
                "window" in text or
                "campaign" in text or
                "kampanya" in text or
                "period" in text
            ):
                campaign_col = col_idx

            if start_col is None and (
                text == "start" or
                "baslangic" in text or
                "başlangıç" in str(value).lower() or
                "start date" in text
            ):
                start_col = col_idx

            if end_col is None and (
                text == "end" or
                "bitis" in text or
                "bitiş" in str(value).lower() or
                "end date" in text
            ):
                end_col = col_idx

        if campaign_col is not None and start_col is not None and end_col is not None:
            header_row = idx
            break

    # Header bulunamazsa eski standart yapıya fallback: 0=campaign, 2=start, 3=end
    if header_row is None:
        header_row = 2 if len(raw) > 2 else 0
        campaign_col = 0
        start_col = 2
        end_col = 3

    data = raw.iloc[header_row + 1:].copy()

    # Bazı dosyalarda header satırı olmadığı halde data 2. satırdan başlıyor olabilir.
    # Eğer fallback data boş kalırsa header row dahil dene.
    if data.empty:
        data = raw.iloc[header_row:].copy()

    campaign = pd.DataFrame({
        "campaign": data.iloc[:, campaign_col] if campaign_col < data.shape[1] else "",
        "start": data.iloc[:, start_col] if start_col < data.shape[1] else pd.NaT,
        "end": data.iloc[:, end_col] if end_col < data.shape[1] else pd.NaT,
    })

    campaign = campaign.dropna(subset=["campaign"])
    campaign["campaign"] = campaign["campaign"].astype(str).str.strip()
    campaign = campaign[campaign["campaign"] != ""]

    campaign["start"] = campaign["start"].apply(excel_serial_to_date)
    campaign["end"] = campaign["end"].apply(excel_serial_to_date)

    campaign = campaign.dropna(subset=["start", "end"])
    campaign = campaign[campaign["end"] >= campaign["start"]]

    return campaign[["campaign", "start", "end"]].reset_index(drop=True)


def read_mapping_file(mapping_file):
    """
    Ürün tipi dosyasını okur.

    Tercih edilen yeni yapı:
    - Ana Ürün sheet: kod listesi
    - Mini Sample sheet: kod listesi
    - ADR sheet: kod listesi
    - kampanya sheet: WINDOW / Start / End

    Eski yapı da desteklenir:
    - ürün tipi sheet: Current Code / Ürün Tipi
    - Adr sheet: Article
    """
    xls = pd.ExcelFile(mapping_file)

    # Yeni, net yapı: ayrı sheetlerden kod okuma
    ana_codes, ana_sheet = read_code_sheet(
        xls,
        ["Ana Ürün", "Ana Urun", "Ana", "Main", "Ana Kodlar"]
    )

    mini_codes, mini_sheet = read_code_sheet(
        xls,
        ["Mini Sample", "Mini", "Sample", "Mini Kodlar", "Sample Kodlar"]
    )

    adr_codes, adr_sheet = read_code_sheet(
        xls,
        ["ADR", "Adr", "ADR Kodlar", "Adr Kodlar"]
    )

    mapping_rows = []

    for code in ana_codes:
        mapping_rows.append({"product_code": code, "product_type": "Ana Ürün"})

    for code in mini_codes:
        mapping_rows.append({"product_code": code, "product_type": "Mini Sample"})

    type_map = pd.DataFrame(mapping_rows)

    # Eğer ayrı Ana Ürün sheet yoksa ama "ürün tipi" sheet varsa, Ana/Mini bilgisini buradan tamamla.
    if any(clean_text(s) == clean_text("ürün tipi") for s in xls.sheet_names):
        product_type_df = pd.read_excel(mapping_file, sheet_name="ürün tipi")
        product_type_df.columns = [str(c).strip() for c in product_type_df.columns]

        code_col_candidates = ["Current Code", "Kod", "Code", "Product Code", "Ürün Kodu", "Urun Kodu"]
        type_col_candidates = ["Ürün Tipi", "Urun Tipi", "Product Type", "Tip"]

        code_col = next((c for c in code_col_candidates if c in product_type_df.columns), None)
        type_col = next((c for c in type_col_candidates if c in product_type_df.columns), None)

        if code_col and type_col:
            product_type_df["product_code"] = product_type_df[code_col].apply(normalize_code)
            product_type_df["product_type"] = product_type_df[type_col].apply(normalize_product_type)

            old_map = (
                product_type_df[["product_code", "product_type"]]
                .dropna()
                .drop_duplicates(subset=["product_code"], keep="first")
            )

            # Ayrı sheetlerden gelen bilgi varsa öncelik ayrı sheetlerde, eksikler ürün tipi sheetinden tamamlanır.
            type_map = pd.concat([old_map, type_map], ignore_index=True)
            type_map = type_map.drop_duplicates(subset=["product_code"], keep="last")

    # ADR sheet eski formatta sadece Article kolonundan da gelebilir; yukarıdaki ilk dolu kolon zaten bunu yakalar.
    # Öncelik ADR'de: bir kod ADR listesinde varsa raporda ADR sayılacak.

    if type_map.empty:
        type_map = pd.DataFrame(columns=["product_code", "product_type"])
    else:
        type_map = type_map.drop_duplicates(subset=["product_code"], keep="last")

    campaign = read_campaign_sheet(xls)

    sheet_info = {
        "Ana Ürün Sheet": ana_sheet or "-",
        "Mini Sample Sheet": mini_sheet or "-",
        "ADR Sheet": adr_sheet or "-",
        "Ana Kod Sayısı": len(ana_codes),
        "Mini Kod Sayısı": len(mini_codes),
        "ADR Kod Sayısı": len(adr_codes),
    }

    return type_map, adr_codes, campaign, sheet_info


def read_supply_file(supply_file):
    """
    Supply dosyasında 'supply' sheetini okur.
    Beklenen yapı:
    Row 2: Libellé- | Calendar day | 20/04/2026 | 27/04/2026 ...
    Row 4+: Product code | Product name | quantities...
    """
    raw = pd.read_excel(supply_file, sheet_name="supply", header=None)

    header_row = 1
    headers = list(raw.iloc[header_row])
    df = raw.iloc[header_row + 2:].copy()
    df.columns = headers

    # İlk iki kolon: ürün kodu ve ürün adı
    code_col = headers[0]
    name_col = headers[1]
    date_cols = headers[2:]

    df = df.rename(columns={code_col: "product_code", name_col: "product_name"})
    df["product_code"] = df["product_code"].apply(normalize_code)

    long_df = df.melt(
        id_vars=["product_code", "product_name"],
        value_vars=date_cols,
        var_name="date",
        value_name="inbound_qty"
    )

    long_df["date"] = long_df["date"].apply(excel_serial_to_date)
    # Supply dosyasındaki tarih, gerçek depo girişinden 1 hafta sonrası olduğu için 7 gün geri çekiyoruz.
    long_df["date"] = long_df["date"] - pd.Timedelta(days=7)
    long_df["inbound_qty"] = pd.to_numeric(long_df["inbound_qty"], errors="coerce").fillna(0)

    long_df = long_df.dropna(subset=["product_code", "date"])
    long_df = long_df[long_df["inbound_qty"] != 0]

    long_df["week_start"] = get_week_start(long_df["date"])

    return long_df


def read_apo_file(apo_file):
    """
    APO Forecast dosyasında 'Weekly ForeCast' sheetini okur.
    Beklenen yapı:
    Row 2: tarih serialleri
    Row 3+: ürün kodu ve haftalık çıkış forecastleri
    """
    raw = pd.read_excel(apo_file, sheet_name="Weekly ForeCast", header=None)

    date_row = 1
    date_values = list(raw.iloc[date_row, 1:])

    df = raw.iloc[date_row + 1:].copy()
    df = df.dropna(how="all")

    code_col = df.columns[0]
    qty_cols = list(df.columns[1:])

    rename_map = {code_col: "product_code"}
    for col, date_value in zip(qty_cols, date_values):
        rename_map[col] = excel_serial_to_date(date_value)

    df = df.rename(columns=rename_map)
    date_cols = [c for c in df.columns if isinstance(c, pd.Timestamp)]

    df["product_code"] = df["product_code"].apply(normalize_code)

    long_df = df.melt(
        id_vars=["product_code"],
        value_vars=date_cols,
        var_name="date",
        value_name="outbound_qty"
    )

    long_df["date"] = pd.to_datetime(long_df["date"], errors="coerce")
    long_df["outbound_qty"] = pd.to_numeric(long_df["outbound_qty"], errors="coerce").fillna(0)

    long_df = long_df.dropna(subset=["product_code", "date"])
    long_df = long_df[long_df["outbound_qty"] != 0]

    long_df["week_start"] = get_week_start(long_df["date"])

    return long_df



def read_ekol_file(ekol_file):
    """
    Ekol depo data dosyasını okur.

    Dosya yapısı:
    - Solda haftalık tablo:
      Depo Yeri | STOK | 13.04 | 20.04 | 27.04 | ...
    - Sağda kapasite özeti:
      Alan | Kapasite | Doluluk | Boş Lokasyon
    """
    if ekol_file is None:
        return None, None

    xls = pd.ExcelFile(ekol_file)
    sheet_name = xls.sheet_names[0]

    raw = pd.read_excel(ekol_file, sheet_name=sheet_name, header=None)
    raw = raw.dropna(how="all")

    if raw.empty:
        return None, None

    # --------------------------------------------------------
    # 1) Haftalık stok tablosu
    # --------------------------------------------------------
    # İlk satır header olarak kullanılır.
    header = list(raw.iloc[0])

    # Sol tablo, ilk boş kolona kadar devam ediyor.
    # Örnek: 0 Depo Yeri, 1 STOK, 2-6 haftalar, 7-8 boş.
    left_cols = []
    for idx, value in enumerate(header):
        if idx >= 2 and pd.isna(value):
            break
        if idx < len(header):
            left_cols.append(idx)

    # Güvenlik: En az ilk 2 kolon + 1 tarih kolonu yoksa None
    ekol_weekly = None
    if len(left_cols) >= 3:
        weekly_raw = raw.iloc[:, left_cols].copy()
        weekly_raw.columns = weekly_raw.iloc[0]
        weekly_raw = weekly_raw.iloc[1:].copy()
        weekly_raw = weekly_raw.dropna(how="all")

        # Kolon adlarını temizle
        new_cols = []
        for col in weekly_raw.columns:
            if pd.isna(col):
                new_cols.append("")
            elif isinstance(col, (int, float, np.integer, np.floating)):
                # 13.04 gibi başlıkları 13.04 olarak göster
                new_cols.append(f"{float(col):.2f}")
            else:
                new_cols.append(str(col).strip())

        weekly_raw.columns = new_cols
        ekol_weekly = weekly_raw.reset_index(drop=True)

    # --------------------------------------------------------
    # 2) Kapasite özeti
    # --------------------------------------------------------
    ekol_capacity = None

    # Header satırında "Alan" olan kolonu bul
    alan_col = None
    for idx, value in enumerate(header):
        if clean_text(value) == "alan":
            alan_col = idx
            break

    if alan_col is not None:
        capacity_raw = raw.iloc[:, alan_col:alan_col + 4].copy()
        capacity_raw.columns = capacity_raw.iloc[0]
        capacity_raw = capacity_raw.iloc[1:].copy()
        capacity_raw = capacity_raw.dropna(how="all")

        # Alan kolonu dolu olanları al
        first_col = capacity_raw.columns[0]
        capacity_raw = capacity_raw[capacity_raw[first_col].notna()]

        ekol_capacity = capacity_raw.reset_index(drop=True)

    return ekol_weekly, ekol_capacity




def format_numeric_dataframe(df):
    """
    Ekranda sayıları virgüllü göstermek için object'e çevrilmiş güvenli tablo döndürür.
    Büyük style kullanımını önlemek için st.dataframe'e düz dataframe veriyoruz.
    """
    if df is None:
        return None

    out = df.copy()

    for col in out.columns:
        numeric = pd.to_numeric(out[col], errors="coerce")
        if numeric.notna().sum() > 0 and numeric.notna().sum() >= max(1, len(out) * 0.4):
            out[col] = numeric.map(lambda x: "" if pd.isna(x) else f"{x:,.0f}")

    return out


def add_product_type(df, type_map, adr_codes):
    """
    Ana/Mini kırılımını oluşturur.
    Önemli: ADR ürünler Ana Ürün içinden çıkarılmaz.
    ADR ayrıca ikinci bir hesap olarak gösterilir ve final paletten düşülür.
    """
    df = df.copy()
    df["product_code"] = df["product_code"].apply(normalize_code)

    df = df.merge(type_map, on="product_code", how="left")
    df["product_type"] = df["product_type"].fillna("Ana Ürün")
    df["product_type"] = df["product_type"].replace({"Merch": "Ana Ürün", "Diğer": "Ana Ürün"})

    df["is_adr"] = df["product_code"].isin(adr_codes)

    # Ana/Mini rapor tipi. ADR burada Ana'nın içinde kalır.
    df["report_type"] = df["product_type"]

    return df


def assign_campaign(week_start, campaign_df):
    if campaign_df.empty or pd.isna(week_start):
        return ""

    week_end = week_start + pd.Timedelta(days=6)

    active = campaign_df[
        (campaign_df["start"] <= week_end) &
        (campaign_df["end"] >= week_start)
    ]

    if active.empty:
        return ""

    return " / ".join(active["campaign"].astype(str).unique())


def safe_divide(a, b):
    if b == 0:
        return 0
    return a / b

def highlight_capacity_and_kpi(df):
    """
    Palet, tır ve kapasite risk alanlarını renklendirir.
    Hem yatay hem dikey tabloda çalışır.
    """
    styles = pd.DataFrame("", index=df.index, columns=df.columns)

    # Yatay tabloda satır adına göre renklendirme
    for row_label in df.index:
        row_text = str(row_label)

        if "Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #d6eaf8;"

        if "ADR Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #fdebd0;"

        if "Total Palet" in row_text:
            styles.loc[row_label, :] = "background-color: #d5f5e3; font-weight: bold;"

        if "Tır Sayısı" in row_text:
            styles.loc[row_label, :] = "background-color: #e8daef; font-weight: bold;"

        if "Kapasite Kullanım %" in row_text:
            for col in df.columns:
                val = pd.to_numeric(pd.Series([df.loc[row_label, col]]), errors="coerce").iloc[0]
                if pd.isna(val):
                    continue
                if val >= 100:
                    styles.loc[row_label, col] = "background-color: #c0392b; color: white; font-weight: bold;"
                elif val >= kritik_esigi:
                    styles.loc[row_label, col] = "background-color: #f5b7b1; font-weight: bold;"
                elif val >= takip_esigi:
                    styles.loc[row_label, col] = "background-color: #f9e79f; font-weight: bold;"
                else:
                    styles.loc[row_label, col] = "background-color: #abebc6; font-weight: bold;"

        if "Haftalık Palet Değişimi" in row_text:
            for col in df.columns:
                val = pd.to_numeric(pd.Series([df.loc[row_label, col]]), errors="coerce").iloc[0]
                if pd.isna(val):
                    continue
                if val > 0:
                    styles.loc[row_label, col] = "background-color: #f8d7da; color: #842029; font-weight: bold;"
                elif val < 0:
                    styles.loc[row_label, col] = "background-color: #d1e7dd; color: #0f5132; font-weight: bold;"

    # Dikey tabloda kolon adına göre renklendirme
    for col in df.columns:
        col_text = str(col)

        if "Palet" in col_text:
            styles[col] = "background-color: #d6eaf8;"

        if "ADR Palet" in col_text:
            styles[col] = "background-color: #fdebd0;"

        if "Total Palet" in col_text:
            styles[col] = "background-color: #d5f5e3; font-weight: bold;"

        if "Tır Sayısı" in col_text:
            styles[col] = "background-color: #e8daef; font-weight: bold;"

        if "Kapasite Kullanım %" in col_text:
            for i, val in enumerate(pd.to_numeric(df[col], errors="coerce")):
                if pd.isna(val):
                    continue
                if val >= 100:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #c0392b; color: white; font-weight: bold;"
                elif val >= kritik_esigi:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #f5b7b1; font-weight: bold;"
                elif val >= takip_esigi:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #f9e79f; font-weight: bold;"
                else:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #abebc6; font-weight: bold;"

        if "Haftalık Palet Değişimi" in col_text:
            for i, val in enumerate(pd.to_numeric(df[col], errors="coerce")):
                if pd.isna(val):
                    continue
                if val > 0:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #f8d7da; color: #842029; font-weight: bold;"
                elif val < 0:
                    styles.iloc[i, df.columns.get_loc(col)] = "background-color: #d1e7dd; color: #0f5132; font-weight: bold;"

    return styles


def highlight_increased_pallet_columns(dataframe, increased_weeks):
    """
    Yatay tabloda Total Palet bir önceki haftaya göre artmışsa,
    ilgili haftanın tüm kolonunu açık kırmızı gösterir.
    """
    styles = pd.DataFrame("", index=dataframe.index, columns=dataframe.columns)

    for col in dataframe.columns:
        week_key = str(col).split("\\n")[0]
        if week_key in increased_weeks:
            styles[col] = "background-color: #f8d7da; color: #842029; font-weight: 600;"

    return styles




def add_months(ts, months):
    """Pandas Timestamp üzerine ay ekler."""
    return pd.Timestamp(ts) + pd.DateOffset(months=months)


def highlight_after_horizon_columns(dataframe, horizon_week_keys):
    """
    5 ay sonrası veri tam olmadığı için ilgili hafta kolonlarını kırmızı/pembe işaretler.
    """
    styles = pd.DataFrame("", index=dataframe.index, columns=dataframe.columns)

    for col in dataframe.columns:
        week_key = str(col).split("\\n")[0]
        if week_key in horizon_week_keys:
            styles[col] = "background-color: #fde2e2; border-left: 4px solid #c0392b;"

    return styles


def safe_format_cell(x):
    """Sayıları virgüllü ve ondalıksız gösterir; metinleri olduğu gibi bırakır."""
    if isinstance(x, (int, float, np.integer, np.floating)) and not pd.isna(x):
        return f"{x:,.0f}"
    return x


def format_excel_workbook(excel_bytes):
    """
    Export edilen Excel dosyasındaki sayısal hücreleri virgüllü ve ondalıksız formatlar.
    """
    bio = BytesIO(excel_bytes)
    wb = load_workbook(bio)

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)
    number_format = '#,##0'

    for ws in wb.worksheets:
        # Header formatı
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sayı formatı
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

        # Kolon genişliği
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 30)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()



# ------------------------------------------------------------
# DEPO EKRANI
# ------------------------------------------------------------
if st.session_state["role"] == "depo":
    show_app_header("Depo operasyon ekranı")

    if st.sidebar.button("Çıkış Yap"):
        st.session_state["role"] = None
        st.rerun()

    history_files = list_saved_reports()

    if not history_files:
        st.warning("Henüz kayıtlı rapor yok. Rapor önce tedarik ekranından kaydedilmelidir.")
        st.stop()

    report_search = st.text_input("Rapor ara", placeholder="Örn. W29, Final, Rev2")
    filtered_history_files = filter_report_names(history_files, report_search)

    if not filtered_history_files:
        st.warning("Aramanızla eşleşen rapor bulunamadı.")
        st.stop()

    selected_history = st.selectbox(
        "Kayıtlı rapor seç",
        options=filtered_history_files
    )

    selected_report_bytes = load_saved_report(selected_history)

    try:
        xls_history = pd.ExcelFile(BytesIO(selected_report_bytes))
    except Exception as e:
        st.error(f"Rapor açılırken hata oluştu: {e}")
        st.stop()

    allowed_sheets = [
        "Depo Haftalık Operasyon",
        "Aylık Giriş Çıkış",
        "Ekol Kapasite Özeti",
        "Ekol Haftalık Stok",
    ]

    available_sheets = [s for s in allowed_sheets if s in xls_history.sheet_names]

    # Eski kayıtlarda Depo Operasyon sheet'i yoksa Veri sheetinden güvenli kolonları çıkar.
    if not available_sheets and "Veri" in xls_history.sheet_names:
        available_sheets = ["Veri"]

    if not available_sheets:
        st.warning("Bu raporda depo ekranında gösterilecek uygun veri bulunamadı. Raporu güncel tedarik ekranından tekrar kaydederseniz haftalık operasyon tablosu görünür.")
        st.stop()

    tabs = st.tabs(available_sheets)

    hidden_keywords = [
        "Palet",
        "Kapasite Kullanım",
        "Kalan Kapasite",
        "Düşülecek",
        "Trend",
        "Tır",
        "Truck",
        "pallet",
        "truck",
    ]

    for tab, sheet_name in zip(tabs, available_sheets):
        with tab:
            df = pd.read_excel(BytesIO(selected_report_bytes), sheet_name=sheet_name)

            if sheet_name == "Depo Haftalık Operasyon":
                st.subheader("Haftalık Operasyon Tablosu")
                st.caption("Haftalar kolonlarda; giriş, çıkış ve stok seviyesi satırlarda gösterilir.")
                df_display = df.copy()

            elif sheet_name == "Veri":
                safe_cols = [
                    "Hafta",
                    "Hafta Başlangıcı",
                    "Kampanya",
                    "Ana Ürün Giriş",
                    "Ana Ürün Çıkış",
                    "Ana Ürün Ekol Stok Seviyesi",
                    "Mini Sample Giriş",
                    "Mini Sample Çıkış",
                    "Mini Sample Ekol Stok Seviyesi",
                    "ADR Giriş",
                    "ADR Çıkış",
                    "ADR Ekol Stok Seviyesi",
                ]
                df = df[[c for c in safe_cols if c in df.columns]]

                if "Hafta" in df.columns:
                    meta_cols = ["Hafta", "Hafta Başlangıcı", "Kampanya"]
                    value_cols = [c for c in df.columns if c not in meta_cols]
                    df_for_pivot = df.copy()
                    df_for_pivot["Hafta Kolonu"] = df_for_pivot.apply(
                        lambda r: f"{r.get('Hafta', '')}\n{r.get('Hafta Başlangıcı', '')}\n{r.get('Kampanya', '')}",
                        axis=1
                    )
                    df_display = df_for_pivot.set_index("Hafta Kolonu")[value_cols].T
                else:
                    df_display = df.copy()

            elif sheet_name not in ["Ekol Kapasite Özeti", "Ekol Haftalık Stok"]:
                visible_cols = [
                    col for col in df.columns
                    if not any(k.lower() in str(col).lower() for k in hidden_keywords)
                ]
                df_display = df[visible_cols]

            else:
                df_display = df.copy()

            # Depo ekranında planning hesaplarını hiçbir şekilde gösterme.
            forbidden_pattern = r"palet|kapasite|trend|tır|truck|düşülecek"

            visible_columns = [
                col for col in df_display.columns
                if not re.search(forbidden_pattern, str(col), flags=re.IGNORECASE)
            ]
            if visible_columns:
                df_display = df_display[visible_columns]

            if len(df_display.index) > 0:
                row_mask = pd.Series(df_display.index.astype(str), index=df_display.index).str.contains(
                    forbidden_pattern, case=False, regex=True, na=False
                )
                df_display = df_display.loc[~row_mask]

            # Sayıları virgüllü ve ondalıksız göster
            df_display = df_display.copy()
            for col in df_display.columns:
                numeric_col = pd.to_numeric(df_display[col], errors="coerce")
                if numeric_col.notna().sum() > 0 and numeric_col.notna().sum() >= max(1, len(df_display) * 0.4):
                    df_display[col] = numeric_col.map(lambda x: "" if pd.isna(x) else f"{x:,.0f}")

            st.dataframe(df_display, use_container_width=True, height=520)

    st.download_button(
        "Seçili Raporu İndir",
        data=selected_report_bytes,
        file_name=selected_history,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.stop()



# TEDARIK_LOGOUT_INSERTED
if st.session_state.get("role") == "tedarik":
    if st.sidebar.button("Çıkış Yap"):
        st.session_state["role"] = None
        st.rerun()

# ------------------------------------------------------------
# UI
# ------------------------------------------------------------
show_app_header("Planning & Warehouse Operations")

st.markdown(
    """
    <div style="
        padding: 12px 16px;
        border: 1px solid #e6e8eb;
        border-radius: 12px;
        margin-bottom: 14px;
        background: #fafbfc;">
        Supply, APO, ürün tipi/kampanya ve Ekol dosyalarını yükleyerek
        haftalık depo hareketlerini ve kapasite görünümünü oluşturun.
    </div>
    """,
    unsafe_allow_html=True
)

st.write(
    "Supply girişlerini, APO çıkış forecastini ve ürün tipi dosyasını yükleyerek "
    "haftalık Ekol stok, palet, ADR palet ve tır hesabını oluşturabilirsiniz."
)

with st.sidebar:
    st.header("Dosyalar")

    supply_file = st.file_uploader("Supply dosyası", type=["xlsx", "xlsm"])
    apo_file = st.file_uploader("APO Forecast dosyası", type=["xlsx", "xlsm"])
    mapping_file = st.file_uploader("Ürün Tipi + Kampanya + ADR dosyası", type=["xlsx", "xlsm"])
    ekol_file = st.file_uploader("Ekol Depo Data Dosyası", type=["xlsx", "xlsm"])

    st.header("Palet Parametreleri")
    ana_palet_ici = st.number_input("Ana Ürün Palet İçi", min_value=1, value=2400, step=50)
    mini_palet_ici = st.number_input("Mini Sample Palet İçi", min_value=1, value=15000, step=100)
    adr_palet_ici = st.number_input("ADR Palet İçi", min_value=1, value=5540, step=10)

    sarf_palet = st.number_input("Haftalık Sarf Palet", min_value=0.0, value=250.0, step=0.5)
    tir_kapasitesi = st.number_input("1 Tır Kaç Palet?", min_value=1, value=40, step=1)

    st.header("Depo Kapasite Parametreleri")
    depo_kapasitesi = st.number_input("Depo Maksimum Palet Kapasitesi", min_value=1.0, value=1100.0, step=100.0)
    takip_esigi = st.number_input("Takip Eşiği (%)", min_value=0.0, max_value=100.0, value=85.0, step=1.0)
    kritik_esigi = st.number_input("Kritik Eşiği (%)", min_value=0.0, max_value=100.0, value=99.0, step=1.0)

    st.header("Rapor Bilgisi")
    report_owner = st.text_input("Raporu Oluşturan", value="Özge")
    report_note = st.text_area("Rapor Notu (Opsiyonel)", placeholder="Örn. Mega Sale devam ediyor.")

    st.header("Başlangıç Stok")
    initial_ana = st.number_input("Başlangıç Ana Ürün Stok", value=0, step=1000)
    initial_mini = st.number_input("Başlangıç Mini Sample Stok", value=0, step=1000)
    initial_adr = st.number_input("Başlangıç ADR Stok", value=0, step=1000)

    calculate = st.button("Raporu Hesapla", type="primary")


pallet_map = {
    "Ana Ürün": ana_palet_ici,
    "Mini Sample": mini_palet_ici,
    "ADR": adr_palet_ici,
}

initial_stock_map = {
    "Ana Ürün": initial_ana,
    "Mini Sample": initial_mini,
    "ADR": initial_adr,
}

mevcut_ana_palet = initial_ana / ana_palet_ici if ana_palet_ici else 0
mevcut_mini_palet = initial_mini / mini_palet_ici if mini_palet_ici else 0
mevcut_adr_palet = initial_adr / adr_palet_ici if adr_palet_ici else 0
mevcut_total_palet = mevcut_ana_palet + mevcut_mini_palet + sarf_palet - mevcut_adr_palet


if calculate:
    if not supply_file or not apo_file or not mapping_file:
        st.error("Lütfen Supply, APO Forecast ve Ürün Tipi dosyalarının üçünü de yükleyin.")
        st.stop()

    with st.spinner("Dosyalar okunuyor ve rapor hazırlanıyor..."):
        type_map, adr_codes, campaign_df, sheet_info = read_mapping_file(mapping_file)

        supply_long = read_supply_file(supply_file)
        apo_long = read_apo_file(apo_file)

        supply_long = add_product_type(supply_long, type_map, adr_codes)
        apo_long = add_product_type(apo_long, type_map, adr_codes)

        # Kontrol için kategori bazlı adet / kod sayısı
        supply_check_base = supply_long.copy()
        supply_check_adr = supply_long[supply_long["is_adr"]].copy()
        supply_check_adr["report_type"] = "ADR"
        supply_check_all = pd.concat([supply_check_base, supply_check_adr], ignore_index=True)

        apo_check_base = apo_long.copy()
        apo_check_adr = apo_long[apo_long["is_adr"]].copy()
        apo_check_adr["report_type"] = "ADR"
        apo_check_all = pd.concat([apo_check_base, apo_check_adr], ignore_index=True)

        supply_category_check = (
            supply_check_all
            .groupby("report_type")
            .agg(
                supply_total_qty=("inbound_qty", "sum"),
                supply_product_count=("product_code", "nunique")
            )
            .reset_index()
        )

        apo_category_check = (
            apo_check_all
            .groupby("report_type")
            .agg(
                apo_total_qty=("outbound_qty", "sum"),
                apo_product_count=("product_code", "nunique")
            )
            .reset_index()
        )

        category_check = pd.merge(
            supply_category_check,
            apo_category_check,
            on="report_type",
            how="outer"
        ).fillna(0)

        # Ana/Mini hesapları: ADR ürünler Ana Ürün içinde kalır.
        inbound_base = (
            supply_long
            .groupby(["week_start", "report_type"], as_index=False)["inbound_qty"]
            .sum()
        )

        outbound_base = (
            apo_long
            .groupby(["week_start", "report_type"], as_index=False)["outbound_qty"]
            .sum()
        )

        # ADR ayrıca gösterilir: Ana Ürün içinden düşülmez, sadece ek satır olarak hesaplanır.
        inbound_adr = (
            supply_long[supply_long["is_adr"]]
            .groupby("week_start", as_index=False)["inbound_qty"]
            .sum()
        )
        inbound_adr["report_type"] = "ADR"

        outbound_adr = (
            apo_long[apo_long["is_adr"]]
            .groupby("week_start", as_index=False)["outbound_qty"]
            .sum()
        )
        outbound_adr["report_type"] = "ADR"

        inbound = pd.concat([inbound_base, inbound_adr], ignore_index=True)
        outbound = pd.concat([outbound_base, outbound_adr], ignore_index=True)

        movement = pd.merge(
            inbound,
            outbound,
            on=["week_start", "report_type"],
            how="outer"
        ).fillna(0)

        # Aylık giriş / çıkış toplamları
        monthly_summary = movement.copy()
        monthly_summary["week_start"] = pd.to_datetime(monthly_summary["week_start"].astype(str), errors="coerce")
        monthly_summary["month"] = monthly_summary["week_start"].dt.strftime("%Y-%m")
        monthly_summary = (
            monthly_summary
            .groupby(["month", "report_type"], as_index=False)[["inbound_qty", "outbound_qty"]]
            .sum()
        )

        monthly_summary = monthly_summary.pivot_table(
            index="month",
            columns="report_type",
            values=["inbound_qty", "outbound_qty"],
            aggfunc="sum",
            fill_value=0
        )

        monthly_summary.columns = [f"{metric}_{rtype}" for metric, rtype in monthly_summary.columns]
        monthly_summary = monthly_summary.reset_index()

        monthly_report = pd.DataFrame()
        monthly_report["Ay"] = monthly_summary["month"]

        for col in [
            "inbound_qty_Ana Ürün", "outbound_qty_Ana Ürün",
            "inbound_qty_Mini Sample", "outbound_qty_Mini Sample",
            "inbound_qty_ADR", "outbound_qty_ADR"
        ]:
            if col not in monthly_summary.columns:
                monthly_summary[col] = 0

        monthly_report["Ana Ürün Giriş"] = monthly_summary["inbound_qty_Ana Ürün"]
        monthly_report["Ana Ürün Çıkış"] = monthly_summary["outbound_qty_Ana Ürün"]
        monthly_report["Mini Sample Giriş"] = monthly_summary["inbound_qty_Mini Sample"]
        monthly_report["Mini Sample Çıkış"] = monthly_summary["outbound_qty_Mini Sample"]
        monthly_report["ADR Giriş"] = monthly_summary["inbound_qty_ADR"]
        monthly_report["ADR Çıkış"] = monthly_summary["outbound_qty_ADR"]

        monthly_numeric_cols = monthly_report.select_dtypes(include=[np.number]).columns
        monthly_report[monthly_numeric_cols] = monthly_report[monthly_numeric_cols].round(0).astype(int)

        # Tüm haftalar x tüm tipler matrisi
        all_weeks = pd.DataFrame({"week_start": sorted(movement["week_start"].dropna().unique())})
        all_types = pd.DataFrame({"report_type": ["Ana Ürün", "Mini Sample", "ADR"]})
        grid = all_weeks.merge(all_types, how="cross")

        movement = grid.merge(movement, on=["week_start", "report_type"], how="left").fillna(0)
        movement = movement.sort_values(["report_type", "week_start"])

        # Kümülatif stok
        stock_rows = []
        for report_type, g in movement.groupby("report_type"):
            current_stock = initial_stock_map.get(report_type, 0)

            for _, row in g.sort_values("week_start").iterrows():
                current_stock = current_stock + row["inbound_qty"] - row["outbound_qty"]

                pallet_inner = pallet_map.get(report_type, ana_palet_ici)
                pallet = safe_divide(current_stock, pallet_inner)

                stock_rows.append({
                    "week_start": row["week_start"],
                    "report_type": report_type,
                    "inbound_qty": row["inbound_qty"],
                    "outbound_qty": row["outbound_qty"],
                    "stock_qty": current_stock,
                    "pallet_inner": pallet_inner,
                    "pallet": pallet,
                })

        detail = pd.DataFrame(stock_rows)

        # Haftalık özet
        weekly = (
            detail
            .pivot_table(
                index="week_start",
                columns="report_type",
                values=["inbound_qty", "outbound_qty", "stock_qty", "pallet"],
                aggfunc="sum"
            )
        )

        weekly.columns = [f"{metric}_{rtype}" for metric, rtype in weekly.columns]
        weekly = weekly.reset_index()
        weekly["week_start"] = pd.to_datetime(weekly["week_start"], errors="coerce")
        weekly = weekly.sort_values("week_start").reset_index(drop=True)

        # Eksik kolonları oluştur
        needed_numeric_cols = [
            "inbound_qty_Ana Ürün", "outbound_qty_Ana Ürün", "stock_qty_Ana Ürün", "pallet_Ana Ürün",
            "inbound_qty_Mini Sample", "outbound_qty_Mini Sample", "stock_qty_Mini Sample", "pallet_Mini Sample",
            "inbound_qty_ADR", "outbound_qty_ADR", "stock_qty_ADR", "pallet_ADR",
        ]

        for col in needed_numeric_cols:
            if col not in weekly.columns:
                weekly[col] = 0

        # Excel görünümüne yakın rapor
        report = pd.DataFrame()
        weekly["week_start"] = pd.to_datetime(weekly["week_start"].astype(str), errors="coerce")
        report["Hafta"] = weekly["week_start"].apply(iso_week_label)
        report["Hafta Başlangıcı"] = weekly["week_start"].dt.strftime("%d.%m.%Y")
        report["Kampanya"] = weekly["week_start"].apply(lambda x: assign_campaign(x, campaign_df))

        report["Ana Ürün Giriş"] = weekly["inbound_qty_Ana Ürün"]
        report["Ana Ürün Çıkış"] = weekly["outbound_qty_Ana Ürün"]
        report["Ana Ürün Ekol Stok Seviyesi"] = weekly["stock_qty_Ana Ürün"]
        report["Ana Ürün Palet"] = weekly["pallet_Ana Ürün"]
        report["Ana Ürün Giriş Paleti"] = report["Ana Ürün Giriş"] / ana_palet_ici

        report["Mini Sample Giriş"] = weekly["inbound_qty_Mini Sample"]
        report["Mini Sample Çıkış"] = weekly["outbound_qty_Mini Sample"]
        report["Mini Sample Ekol Stok Seviyesi"] = weekly["stock_qty_Mini Sample"]
        report["Mini Sample Palet"] = weekly["pallet_Mini Sample"]
        report["Mini Sample Giriş Paleti"] = report["Mini Sample Giriş"] / mini_palet_ici

        report["ADR Giriş"] = weekly["inbound_qty_ADR"]
        report["ADR Çıkış"] = weekly["outbound_qty_ADR"]
        report["ADR Ekol Stok Seviyesi"] = weekly["stock_qty_ADR"]
        report["ADR Palet"] = weekly["pallet_ADR"]

        # ADR palet ayrı gösterilir ve final total paletten düşülür.
        report["Sarf Palet"] = sarf_palet
        report["ADR Düşülecek Palet"] = report["ADR Palet"]
        report["Total Palet"] = (
            report["Ana Ürün Palet"] +
            report["Mini Sample Palet"] +
            report["Sarf Palet"] -
            report["ADR Düşülecek Palet"]
        )

        # Tır sayısı hesabı: sadece o haftaki girişlerin Ana Ürün + Mini Sample giriş paleti üzerinden yapılır.
        report["Tır Sayısı"] = (
            report["Ana Ürün Giriş Paleti"] +
            report["Mini Sample Giriş Paleti"]
        ) / tir_kapasitesi

        # Depo kapasite kontrol alanları
        report["Kapasite Kullanım %"] = report["Total Palet"] / depo_kapasitesi * 100
        report["Kalan Kapasite Palet"] = depo_kapasitesi - report["Total Palet"]
        report["Haftalık Palet Değişimi"] = report["Total Palet"].diff().fillna(0)

        def capacity_status(value):
            if value >= 100:
                return "Kapasite Aşımı"
            if value >= kritik_esigi:
                return "Kritik"
            if value >= takip_esigi:
                return "Takip"
            return "Güvenli"

        report["Kapasite Durumu"] = report["Kapasite Kullanım %"].apply(capacity_status)

        report["Palet Trend"] = np.where(
            report["Haftalık Palet Değişimi"] > 0,
            "Artış",
            np.where(report["Haftalık Palet Değişimi"] < 0, "Azalış", "Sabit")
        )

        numeric_cols = report.select_dtypes(include=[np.number]).columns
        report[numeric_cols] = report[numeric_cols].round(0).astype(int)

        weekly = report


    # Yeni sekmeli sonuç alanı
    tab_dashboard, tab_weekly, tab_current, tab_monthly, tab_ekol = st.tabs([
        "🏠 Dashboard",
        "📅 Haftalık Özet",
        "📦 Mevcut Hafta",
        "📈 Aylık Giriş / Çıkış",
        "🚚 Ekol",
    ])

    with tab_dashboard:
        st.success("Rapor hazır.")
        info1, info2, info3 = st.columns(3)
        info1.info(f"Oluşturan: {report_owner or '-'}")
        info2.info(f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        info3.info(f"Not: {report_note or '-'}")

        # KPI'lar ilk haftanın değerlerini gösterir.
        first_total = weekly["Total Palet"].iloc[0]
        first_capacity = weekly["Kapasite Kullanım %"].iloc[0]
        first_remaining = weekly["Kalan Kapasite Palet"].iloc[0]
        first_status = weekly["Kapasite Durumu"].iloc[0]
        first_week = weekly["Hafta"].iloc[0]

        # Peak hafta sadece ilk 5 ay içinde aranır.
        weekly_for_peak = weekly.copy()
        weekly_for_peak["_date"] = pd.to_datetime(weekly_for_peak["Hafta Başlangıcı"], dayfirst=True, errors="coerce")
        first_date = weekly_for_peak["_date"].min()
        horizon_date = add_months(first_date, 5)
        first_5_months = weekly_for_peak[weekly_for_peak["_date"] < horizon_date].copy()

        if first_5_months.empty:
            first_5_months = weekly_for_peak.copy()

        peak_idx = first_5_months["Total Palet"].idxmax()
        peak_week = first_5_months.loc[peak_idx, "Hafta"]
        peak_pallet = first_5_months.loc[peak_idx, "Total Palet"]
        increasing_week_count = int((first_5_months["Palet Trend"] == "Artış").sum())

        c1, c2, c3, c4, c5 = st.columns(5)

        with c1:
            st.metric("İlk Hafta Total Palet", f"{first_total:,.0f}", first_week)
        with c2:
            st.metric("İlk Hafta Kapasite %", f"{first_capacity:,.0f}%")
        with c3:
            st.metric("İlk Hafta Kalan Kapasite", f"{first_remaining:,.0f} palet")
        with c4:
            st.metric("Peak Hafta / İlk 5 Ay", f"{peak_week}", f"{peak_pallet:,.0f} palet")
        with c5:
            st.metric("Artan Hafta Sayısı / İlk 5 Ay", f"{increasing_week_count}")

        if first_status in ["Kritik", "Kapasite Aşımı"]:
            st.error(f"İlk hafta kapasite durumu: {first_status}")
        elif first_status == "Takip":
            st.warning(f"İlk hafta kapasite durumu: {first_status}")
        else:
            st.success(f"İlk hafta kapasite durumu: {first_status}")

    with tab_weekly:
        st.subheader("Yatay Haftalık Görünüm / Excel Formatı")

        pivot_rows = [
            "Ana Ürün Giriş",
            "Ana Ürün Çıkış",
            "Ana Ürün Ekol Stok Seviyesi",
            "Ana Ürün Palet",

            "Mini Sample Giriş",
            "Mini Sample Çıkış",
            "Mini Sample Ekol Stok Seviyesi",
            "Mini Sample Palet",

            "ADR Giriş",
            "ADR Çıkış",
            "ADR Ekol Stok Seviyesi",
            "ADR Palet",

            "Sarf Palet",
            "ADR Düşülecek Palet",
            "Total Palet",
            "Kapasite Kullanım %",
            "Kalan Kapasite Palet",
            "Kapasite Durumu",
            "Palet Trend",
            "Tır Sayısı",
        ]

        horizontal = weekly.set_index("Hafta")[pivot_rows].T

        weekly_date_map = weekly.set_index("Hafta")["Hafta Başlangıcı"]
        campaign_row = weekly.set_index("Hafta")["Kampanya"]

        horizontal.columns = [
            f"{week}\n{weekly_date_map.loc[week]}\n{campaign_row.loc[week] if campaign_row.loc[week] else ''}"
            for week in horizontal.columns
        ]

        # Total Palet önceki haftaya göre artıyorsa o haftayı kritik kırmızı yap.
        weekly_sorted = weekly.copy()
        weekly_sorted["Total Palet Artış"] = weekly_sorted["Total Palet"].diff()
        increased_weeks = weekly_sorted.loc[
            weekly_sorted["Total Palet Artış"] > 0, "Hafta"
        ].astype(str).tolist()

        # İlk 5 aydan sonraki haftaları kırmızı belirteçle işaretle.
        weekly_horizon = weekly.copy()
        weekly_horizon["_date"] = pd.to_datetime(weekly_horizon["Hafta Başlangıcı"], dayfirst=True, errors="coerce")
        first_week_date = weekly_horizon["_date"].min()
        horizon_limit = add_months(first_week_date, 5)
        after_horizon_weeks = weekly_horizon.loc[
            weekly_horizon["_date"] >= horizon_limit, "Hafta"
        ].astype(str).tolist()

        st.caption("Ekran performansı için tabloda ilk 5 ay gösterilir. Excel çıktısında tüm haftalar yer alır.")

        display_cols = [
            col for col in horizontal.columns
            if str(col).split("\\n")[0] not in after_horizon_weeks
        ]
        horizontal_display = horizontal[display_cols] if display_cols else horizontal

        st.dataframe(format_numeric_dataframe(horizontal_display), use_container_width=True, height=520)

        with st.expander("Okunan Kampanya Takvimi"):
            if campaign_df.empty:
                st.warning("Kampanya takvimi okunamadı. Kampanya sheetinde kampanya adı, start ve end tarihleri olduğundan emin olun.")
            else:
                campaign_view = campaign_df.copy()
                campaign_view["start"] = campaign_view["start"].dt.strftime("%d.%m.%Y")
                campaign_view["end"] = campaign_view["end"].dt.strftime("%d.%m.%Y")
                st.dataframe(campaign_view, use_container_width=True)

    with tab_current:
        st.subheader("Mevcut Hafta Palet Tablosu")

        mevcut_hafta_report = pd.DataFrame({
            "Kategori": ["Ana Ürün", "Mini Sample", "ADR", "Sarf", "Total"],
            "Başlangıç Stok": [initial_ana, initial_mini, initial_adr, np.nan, np.nan],
            "Palet İçi": [ana_palet_ici, mini_palet_ici, adr_palet_ici, np.nan, np.nan],
            "Mevcut Hafta Palet": [
                mevcut_ana_palet,
                mevcut_mini_palet,
                mevcut_adr_palet,
                sarf_palet,
                mevcut_total_palet
            ]
        })

        for col in ["Başlangıç Stok", "Palet İçi", "Mevcut Hafta Palet"]:
            mevcut_hafta_report[col] = pd.to_numeric(mevcut_hafta_report[col], errors="coerce").round(0)

        st.dataframe(
            mevcut_hafta_report.style.format({
                "Başlangıç Stok": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
                "Palet İçi": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
                "Mevcut Hafta Palet": lambda x: "" if pd.isna(x) else f"{x:,.0f}",
            }),
            use_container_width=True
        )

    with tab_monthly:
        st.subheader("Aylık Giriş / Çıkış Toplamları")

        monthly_horizontal = monthly_report.set_index("Ay").T
        st.dataframe(format_numeric_dataframe(monthly_horizontal), use_container_width=True, height=360)

    with tab_ekol:
        # Ekol Depo Data
        ekol_weekly = None
        ekol_capacity = None

        if ekol_file is not None:
            st.subheader("Ekol Depo Data")

            try:
                ekol_weekly, ekol_capacity = read_ekol_file(ekol_file)

                if ekol_capacity is not None:
                    st.write("Ekol Kapasite Özeti")
                    st.dataframe(
                        format_numeric_dataframe(ekol_capacity),
                        use_container_width=True,
                        height=220
                    )

                if ekol_weekly is not None:
                    st.write("Ekol Haftalık Stok / Doluluk Tablosu")
                    st.dataframe(
                        format_numeric_dataframe(ekol_weekly),
                        use_container_width=True,
                        height=420
                    )

                if ekol_weekly is None and ekol_capacity is None:
                    st.warning("Ekol dosyası okundu ancak uygun tablo bulunamadı.")

            except Exception as e:
                st.error(f"Ekol dosyası okunurken hata oluştu: {e}")
    # Depo ekranı için haftalık yatay operasyon tablosu
    depo_operasyon_cols = [
        "Hafta",
        "Hafta Başlangıcı",
        "Kampanya",
        "Ana Ürün Giriş",
        "Ana Ürün Çıkış",
        "Ana Ürün Ekol Stok Seviyesi",
        "Mini Sample Giriş",
        "Mini Sample Çıkış",
        "Mini Sample Ekol Stok Seviyesi",
        "ADR Giriş",
        "ADR Çıkış",
        "ADR Ekol Stok Seviyesi",
    ]

    depo_operasyon = weekly[[c for c in depo_operasyon_cols if c in weekly.columns]].copy()

    if not depo_operasyon.empty and "Hafta" in depo_operasyon.columns:
        meta_cols = ["Hafta", "Hafta Başlangıcı", "Kampanya"]
        value_cols = [c for c in depo_operasyon.columns if c not in meta_cols]

        depo_operasyon["Hafta Kolonu"] = depo_operasyon.apply(
            lambda r: f"{r.get('Hafta', '')}\n{r.get('Hafta Başlangıcı', '')}\n{r.get('Kampanya', '')}",
            axis=1
        )

        depo_yatay_operasyon = depo_operasyon.set_index("Hafta Kolonu")[value_cols].T
    else:
        depo_yatay_operasyon = pd.DataFrame()

    report_metadata = pd.DataFrame([{
        "Rapor Adı": "YR Logistics Dashboard",
        "Oluşturan": report_owner,
        "Oluşturma Tarihi": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "Rapor Notu": report_note,
        "Uygulama Versiyonu": APP_VERSION,
    }])

    # Excel export
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        horizontal.to_excel(writer, sheet_name="Yatay Özet")
        depo_yatay_operasyon.to_excel(writer, sheet_name="Depo Haftalık Operasyon")
        depo_operasyon.to_excel(writer, sheet_name="Depo Operasyon", index=False)
        mevcut_hafta_report.to_excel(writer, sheet_name="Mevcut Hafta Palet", index=False)
        monthly_horizontal.to_excel(writer, sheet_name="Aylık Giriş Çıkış")
        weekly.to_excel(writer, sheet_name="Veri", index=False)
        pd.DataFrame([sheet_info]).to_excel(writer, sheet_name="Okunan Sheet Bilgisi", index=False)
        category_check.to_excel(writer, sheet_name="Kategori Kontrol", index=False)
        campaign_df.to_excel(writer, sheet_name="Kampanya", index=False)
        report_metadata.to_excel(writer, sheet_name="Rapor Bilgisi", index=False)

        if ekol_file is not None:
            if ekol_capacity is not None:
                ekol_capacity.to_excel(writer, sheet_name="Ekol Kapasite Özeti", index=False)

            if ekol_weekly is not None:
                ekol_weekly.to_excel(writer, sheet_name="Ekol Haftalık Stok", index=False)

    # Son oluşturulan raporu hafızada tut. Excel kaydında sayılar virgüllü görünür.
    formatted_report_bytes = format_excel_workbook(output.getvalue())
    st.session_state["last_report_bytes"] = formatted_report_bytes
    st.session_state["last_report_default_name"] = f"YR_Logistics_{datetime.now().strftime('%Y%m%d')}"

    st.download_button(
        label="Excel Raporu İndir",
        data=st.session_state["last_report_bytes"],
        file_name="depo_giris_cikis_dashboard_raporu.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("Soldaki panelden 3 dosyayı yükleyip parametreleri belirledikten sonra 'Raporu Hesapla' butonuna basın.")



# ------------------------------------------------------------
# RAPORU GEÇMİŞE KAYDET
# ------------------------------------------------------------
st.divider()
st.subheader("💾 Raporu Geçmişe Kaydet")

if "last_report_bytes" in st.session_state:
    report_save_name = st.text_input(
        "Geçmişe kaydetme adı",
        value=st.session_state.get("last_report_default_name", f"YR_Logistics_{datetime.now().strftime('%Y%m%d')}")
    )

    if st.button("Geçmişe Kaydet"):
        clean_name = "".join(
            ch if ch.isalnum() or ch in [" ", "_", "-"] else "_"
            for ch in report_save_name.strip()
        )

        if not clean_name:
            clean_name = f"YR_Logistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        history_filename = f"{clean_name}.xlsx"

        try:
            saved_filename = save_report_permanently(
                history_filename,
                st.session_state["last_report_bytes"]
            )
            st.success(f"Rapor geçmişe kalıcı olarak kaydedildi: {saved_filename}")
        except Exception as e:
            st.error(f"Rapor kaydedilirken hata oluştu: {e}")
else:
    st.info("Önce raporu hesaplayın. Rapor hesaplandıktan sonra burada geçmişe kaydedebilirsiniz.")

# ------------------------------------------------------------
# GEÇMİŞ RAPORLAR
# ------------------------------------------------------------
st.divider()
st.subheader("🗂️ Geçmiş Raporlar")

history_files = list_saved_reports()

if history_files:
    history_search = st.text_input("Geçmiş raporlarda ara", placeholder="Örn. W29, Final, Rev2")
    filtered_history_files = filter_report_names(history_files, history_search)

    if not filtered_history_files:
        st.warning("Aramanızla eşleşen geçmiş rapor bulunamadı.")
        st.stop()

    selected_history = st.selectbox(
        "Geçmiş rapor seç",
        options=filtered_history_files
    )

    selected_report_bytes = load_saved_report(selected_history)

    st.download_button(
        "Seçili Geçmiş Raporu İndir",
        data=selected_report_bytes,
        file_name=selected_history,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("Seçili Geçmiş Raporu Aç"):
        xls_history = pd.ExcelFile(BytesIO(selected_report_bytes))
        tabs = st.tabs(xls_history.sheet_names)

        for tab, sheet_name in zip(tabs, xls_history.sheet_names):
            with tab:
                old_df = pd.read_excel(BytesIO(selected_report_bytes), sheet_name=sheet_name)
                st.dataframe(format_numeric_dataframe(old_df), use_container_width=True)

    st.caption(f"Toplam kayıtlı rapor sayısı: {len(history_files)}")
else:
    st.info("Henüz geçmiş rapor yok. Raporu oluşturduktan sonra 'Geçmişe Kaydet' butonuna basarsan burada görünür.")


st.sidebar.caption(f"{APP_NAME} · v{APP_VERSION} · {st.session_state.get('role', '-').title()}")
